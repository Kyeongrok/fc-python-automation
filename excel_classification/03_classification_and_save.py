from datetime import datetime

import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Font, Alignment


# pd.set_option('display.max_columns', None)

class ClassificationExcel:

    path = ''

    def __init__(self, order_xlsx_filename, partner_info_xlsx_filename, path='result'):
        # 주문목록
        df = pd.read_excel(order_xlsx_filename)
        df = df.rename(columns=df.iloc[1])
        df = df.drop([df.index[0], df.index[1]])
        df = df.reset_index(drop=True)
        self.order_list = df
        self.path = path

        # 파트너목록
        df_partners = pd.read_excel(partner_info_xlsx_filename)

        self.brands = df_partners['브랜드'].to_list()
        self.partners = df_partners['업체명'].to_list()


    def classify(self):

        for i, row in self.order_list.iterrows():
            brand_name = ''
            partner_name = ''
            for j in range(len(self.brands)):
                if self.brands[j] in row['상품명']:
                    brand_name = self.brands[j]
                    partner_name = self.partners[j]
                    break
            # print(f'{row["상품명"]} 은 {brand_name} 브랜드 입니다. {j}번째')
            # print(f'업체명:{partner_name}')

            if partner_name != '':
                df_filtered = self.order_list[self.order_list['상품명'].str.contains(brand_name)]
                df_filtered.to_excel(f'{self.path}/[패스트몰] {partner_name}.xlsx')
            else:
                print('없는 brand name:', brand_name, row['상품명'])

    def set_count(self):
        file_name = '20221113/[패스트몰] 다온마켓.xlsx'
        wb = load_workbook(file_name)
        ws = wb.active
        print('value:', ws['B1'].value)
        print('value:', ws['B2'].value)

        # 개수 세기
        row_cnt = ws.max_row - 1
        print('cnt:', row_cnt)

        # 열 삽입
        ws.insert_rows(1)
        ws.insert_rows(1)

        now_day = datetime.now().strftime('%Y-%m-%d')

        # A1
        ws['A1'] = f'발송요청내역 [총 {row_cnt}건] {now_day}'
        ws['A1'].font = Font(size=11, bold=True)
        ws.merge_cells('A1:U1')
        ws['A1'].alignment = Alignment(horizontal='left')

        wb.save(file_name)





if __name__ == '__main__':
    ce = ClassificationExcel('주문목록20221112.xlsx', '파트너목록.xlsx', '20221113')
    # ce.classify()
    ce.set_count()
