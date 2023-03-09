import os
import smtplib
from email.mime.text import MIMEText

from openpyxl.reader.excel import load_workbook


class EmailSender:
    email_addr = None
    password = None
    smtp_server_map = {
        'gmail.com':'smtp.gmail.com',
        'naver.com':'smtp.naver.com'
    }
    smtp_server = None

    def __init__(self, email_addr, password):
        print('생성자')
        self.email_addr = email_addr
        self.password = password
        self.smtp_server = self.smtp_server_map[email_addr.split('@')[1]] # fc.krkim@gmail.com
        print(self.smtp_server)

    def send_email(self, msg, from_addr, to_addr, subject):
        """
        :param msg: 보낼 메세지
        :param from_addr: 보내는 사람
        :param to_addr: 받는 사람
        :return:
        """
        with smtplib.SMTP(self.smtp_server, 587) as smtp:
            msg = MIMEText(msg)
            msg['From'] = from_addr
            msg['To'] = to_addr
            msg['Subject'] = subject
            print(msg.as_string())

            smtp.starttls()
            smtp.login(self.email_addr, self.password)
            smtp.sendmail(from_addr=from_addr, to_addrs=to_addr, msg=msg.as_string())
            smtp.quit()
        print(f'to_addr:{to_addr}로 이메일 전송이 완료 되었습니다.')

    def send_all_emails(self, filename):
        print(f'{filename}에 있는 이메일과 내용을 이용해 메일을 보냅니다.')
        wb = load_workbook(filename)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):
            if row[0].value != None:
                print(row[0].value, row[1].value, row[2].value)
                self.send_email(row[2].value,
                              from_addr=self.email_addr,
                              to_addr=row[0].value, subject=row[1].value)

if __name__ == '__main__':
    # es = EmailSender('fc.krkim@gmail.com', os.getenv('MY_GMAIL_PASSWORD'))
    es = EmailSender('oceanfog@naver.com', os.getenv('MY_NAVER_PASSWORD'))
    es.send_all_emails('이메일리스트.xlsx')