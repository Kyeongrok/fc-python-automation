from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook


class WeeklyWorkPlan:
    wb = None
    ws = None
    start_date = '2022-09-01'
    manager = "매니저 이름을 입력 해주세요"
    date_list = []
    days_of_week = []

    def __init__(self, start_date, manager, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager

        # 날짜 생성
        self.set_date()
        self.set_title()
        self.set_table()

    def save(self, filename):
        self.wb.save(filename)
        print('엑셀파일 생성 완료')

    def set_date(self, days=6):
        # start_date + 6일
        end_date = datetime.strptime(self.start_date, '%Y-%m-%d') + timedelta(days=days)
        # [5 6 7 8 9 10 11]
        week = pd.date_range(start=self.start_date, end=end_date.strftime('%Y-%m-%d'))
        self.date_list = week.strftime('%Y-%m-%d').to_list()
        self.days_of_week = week.strftime('%A').to_list()

        print('end_date:', end_date)
        print('week:', week)
        print('date_list:', self.date_list)
        print('days_of_week:', self.days_of_week)

    def set_title(self):
        ws = self.ws
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = '주간업무계획표'
        start_date = self.date_list[0]
        end_date = self.date_list[-1]
        ws['B6'] = f'({start_date} ~ {end_date})'

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print('타이틀 생성 완료')

    def set_table(self):
        ws = self.ws
        ws['B8'] = '날짜'
        col_names = ['날짜', '요일', '시간', '일정', '비고']

        # 컬럼명
        for i in range(5):
            # print('i:', i, i + 2)
            ws.cell(row=8, column=i + 2).value = col_names[i]

        # 날짜
        for i in range(len(self.date_list)):
            # print('i:', i, i + 2)
            ws.cell(row=9 + i, column=2).value = self.date_list[i]

        # 요일
        for i in range(len(self.days_of_week)):
            ws.cell(row=9 + i, column=3).value = self.days_of_week[i]

        # 행 삽입하기
        ws.insert_rows(10, 4)
        ws.insert_rows(15, 4)
        ws.insert_rows(20, 4)
        ws.insert_rows(25, 4)
        ws.insert_rows(30, 4)
        ws.insert_rows(35, 4)

        # 셀 병합하기
        ws.merge_cells('B9:B13') # 날짜



if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2022-09-05', '김패캠')
    wwp.save('주간업무계획표.xlsx')
