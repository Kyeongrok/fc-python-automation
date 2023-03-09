from datetime import datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


class WeeklyWorkPlan:
    wb = None
    ws = None
    start_date = '2022-09-01'
    manager = "매니저 이름을 입력 해주세요"
    date_list = []
    days_of_week = []

    def __init__(self, start_date, manager, days=5, sheet_no=0):
        self.wb = Workbook()
        self.ws = self.wb.worksheets[sheet_no]
        self.start_date = start_date
        self.manager = manager

        # 날짜 생성
        self.set_date(days=days)
        self.set_title()
        self.set_table()
        self.set_style()

    def save(self, filename):
        self.wb.save(filename)
        print('엑셀파일 생성 완료')

    def set_date(self, days=6):
        # start_date + 6일
        end_date = datetime.strptime(self.start_date, '%Y-%m-%d') + timedelta(days=days)
        # [5 6 7 8 9 10 11]
        week = pd.date_range(start=self.start_date, end=end_date.strftime('%Y-%m-%d'))
        self.date_list = week.strftime('%Y-%m-%d').to_list()
        self.days_of_week = week.strftime('%A').to_list()

        print('end_date:', end_date)
        print('week:', week)
        print('date_list:', self.date_list)
        print('days_of_week:', self.days_of_week)

    def set_title(self):
        ws = self.ws
        ws['B2'] = '담당자'
        ws['C2'] = self.manager
        ws['B3'] = '시작일'
        ws['C3'] = self.start_date

        # 제목
        ws['B5'] = '주간업무계획표'
        start_date = self.date_list[0]
        end_date = self.date_list[-1]
        ws['B6'] = f'({start_date} ~ {end_date})'

        # 셀병합
        ws.merge_cells('B5:F5')
        ws.merge_cells('B6:F6')

        print('타이틀 생성 완료')

    def set_table(self):
        ws = self.ws
        ws['B8'] = '날짜'
        col_names = ['날짜', '요일', '시간', '일정', '비고']

        # 컬럼명
        for i in range(5):
            # print('i:', i, i + 2)
            ws.cell(row=8, column=i + 2).value = col_names[i]

        # 날짜, 요일
        for i in range(len(self.date_list)):
            # print('i:', i, i + 2)
            ws.cell(row=9 + (i * 5), column=2).value = self.date_list[i]
            ws.cell(row=9 + (i * 5), column=3).value = self.days_of_week[i]

            # 셀 병합하기
            ws.merge_cells(f'B{9 + i * 5}:B{13 + i * 5}') # 날짜
            ws.merge_cells(f'C{9 + i * 5}:C{13 + i * 5}') # 요일
            ws.merge_cells(f'F{9 + i * 5}:F{13 + i * 5}') # 비고

    def set_style(self):
        ws = self.ws

        # A열 너비
        ws.column_dimensions['A'].width = 5

        # 열 제목 B C D E F
        for i in range(2, 7):
            # print('get_column_letter:', i, get_column_letter(i))
            ws.column_dimensions[get_column_letter(i)].width = 15
            ws[f'{get_column_letter(i)}8'].font = Font(name='맑은 고딕', bold=True)
            ws[f'{get_column_letter(i)}8'].alignment = Alignment(horizontal='center', vertical='center')

        # E열 너비
        ws.column_dimensions['E'].width = 40

        # 제목 글꼴, 사이즈
        ws['B5'].font = Font(name='맑은 고딕', size=28, bold=True)


        # 가운데 정렬
        ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['B6'].alignment = Alignment(horizontal='center', vertical='center')

        # 날짜 요일 가운데 정렬
        for i in range(9, 40, 5):
            ws[f'B{i}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'C{i}'].alignment = Alignment(horizontal='center', vertical='center')


if __name__ == '__main__':
    wwp = WeeklyWorkPlan('2022-09-04', '김패캠', days=5)
    wwp.save('주간업무계획표.xlsx')
